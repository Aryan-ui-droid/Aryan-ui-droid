import pandas as pd

class ReportGenerator:

    def __init__(self, engine, university, department, semester, academic_year, class_name="", term_key=None, top_n=3):
        self.engine = engine
        self.df = engine.df

        self.university = university
        self.department = department
        self.semester = semester
        self.academic_year = academic_year
        self.class_name = class_name
        self.term_key = term_key
        self.top_n = max(1, int(top_n))
    def add_header(self, ws, max_col=6):
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        # Ensure header has enough width for long department names,
        # even on sheets with few data columns.
        min_cols = 7
        end_col = get_column_letter(max(1, max_col, min_cols))
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dept_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # University
        ws.merge_cells(f"A1:{end_col}1")
        ws["A1"] = self.university
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws["A1"].alignment = header_alignment

        # Department
        ws.merge_cells(f"A2:{end_col}2")
        ws["A2"] = f"Department of {self.department}"
        ws["A2"].font = Font(name="Times New Roman", bold=True)
        ws["A2"].alignment = dept_alignment
        ws.row_dimensions[2].height = 36

        # Title
        ws.merge_cells(f"A3:{end_col}3")
        ws["A3"] = f"RESULT ANALYSIS REPORT - SEMESTER {self.semester}"
        ws["A3"].font = Font(name="Times New Roman", bold=True)
        ws["A3"].alignment = header_alignment

        # Academic Year
        ws.merge_cells(f"A4:{end_col}4")
        ws["A4"] = f"Academic Year: {self.academic_year}"
        ws["A4"].font = Font(name="Times New Roman", bold=True)
        ws["A4"].alignment = header_alignment

    def style_table(self, ws, df, start_row=7):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        header_font = Font(name="Times New Roman", bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        max_col = len(df.columns)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[start_row].height = 24

        for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + len(df), min_col=1, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Times New Roman")
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for val in df[col_name].astype(str).tolist():
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    def apply_pass_percent_fill(self, ws, start_row, row_count, col_idx):
        from openpyxl.styles import PatternFill

        green_fill = PatternFill("solid", fgColor="93C47D")
        yellow_fill = PatternFill("solid", fgColor="FFE599")
        red_fill = PatternFill("solid", fgColor="F4CCCC")

        for row in range(start_row, start_row + row_count):
            cell = ws.cell(row=row, column=col_idx)
            try:
                value = float(str(cell.value).replace("%", "").strip())
            except Exception:
                continue
            if value >= 90:
                cell.fill = green_fill
            elif value >= 75:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

    def add_interpretation_sheet(self, writer, subject_df):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = writer.book
        ws = wb.create_sheet("Interpretation")

        ws.merge_cells("A1:G1")
        ws["A1"] = self.university
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Department: {self.department}"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.row_dimensions[2].height = 24

        ws.merge_cells("A3:G3")
        ws["A3"] = f"SUBJECT WISE RESULT ANALYSIS - {self.semester}"
        ws["A3"].font = Font(bold=True)
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:G4")
        ws["A4"] = f"Academic year: {self.academic_year}"
        ws["A4"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A5:D5")
        ws["A5"] = "Exam Month & Year: -"
        ws["A5"].alignment = Alignment(horizontal="left")

        ws.merge_cells("E5:G5")
        ws["E5"] = f"Sem: {self.semester}"
        ws["E5"].alignment = Alignment(horizontal="right")

        headers = [
            "Sr No",
            "Subject",
            "Name of Faculty",
            "Result",
            "Analysis & Interpretation",
            "Corrective Measures",
            "Sign of Faculty",
        ]
        header_row = 7
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        if subject_df is None or subject_df.empty:
            return

        green_fill = PatternFill("solid", fgColor="93C47D")
        yellow_fill = PatternFill("solid", fgColor="FFE599")
        red_fill = PatternFill("solid", fgColor="F4CCCC")

        for idx, row in subject_df.reset_index(drop=True).iterrows():
            sr_no = idx + 1
            subject_name = str(row.get("subject_name", "")).strip() or f"Subject {sr_no}"
            course_code = str(row.get("course_code", "")).strip()
            subject_label = subject_name if not course_code else f"{subject_name} ({course_code})"
            pass_percent = float(row.get("pass_percent", 0.0))
            failed = int(row.get("failed", 0))

            if pass_percent >= 90:
                analysis = "Good performance"
                result_fill = green_fill
            elif pass_percent >= 75:
                analysis = "Average performance"
                result_fill = yellow_fill
            else:
                analysis = "Needs improvement"
                result_fill = red_fill

            corrective = ""
            if failed > 0:
                corrective = "Counseling is needed for students who have failed"

            data_row = header_row + idx + 1
            values = [
                sr_no,
                subject_label,
                "",
                f"{pass_percent:.2f}%",
                analysis,
                corrective,
                "",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if col == 4:
                    cell.fill = result_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 26
        ws.column_dimensions["F"].width = 34
        ws.column_dimensions["G"].width = 18

    def add_grade_sheet(self, writer):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = writer.book
        ws = wb.create_sheet("Grade Sheet")

        max_col = 21  # A to U
        end_col = get_column_letter(max_col)

        title_font = Font(name="Times New Roman", size=12, bold=True)
        header_font = Font(name="Times New Roman", bold=True)
        yellow_fill = PatternFill("solid", fgColor="FFF200")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        def apply_border_range(cell_range):
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border

        # Header block
        ws.merge_cells(f"A1:{end_col}1")
        ws["A1"] = self.university
        ws["A1"].font = title_font
        ws["A1"].alignment = center

        ws.merge_cells(f"A2:{end_col}2")
        ws["A2"] = f"DEPARTMENT: {self.department}"
        ws["A2"].font = header_font
        ws["A2"].alignment = center

        ws.merge_cells(f"A3:{end_col}3")
        class_label = f"{self.class_name}".strip()
        sem_label = f"{self.semester}".strip()
        ws["A3"] = f"SUBJECT WISE RESULT ANALYSIS {class_label} - {sem_label}".strip()
        ws["A3"].font = header_font
        ws["A3"].alignment = center

        ws.merge_cells(f"A4:{end_col}4")
        ws["A4"] = f"Academic year : {self.academic_year}"
        ws["A4"].font = Font(name="Times New Roman", bold=True)
        ws["A4"].alignment = center

        # Left block + right block
        ws.merge_cells("A6:F6")
        ws["A6"] = f"CLASS : {class_label}".strip()
        ws["A6"].font = header_font
        ws["A6"].alignment = left

        ws.merge_cells("L6:S6")
        ws["L6"] = "RESULT AFTER REM"
        ws["L6"].font = header_font
        ws["L6"].alignment = center

        overview = self.engine.get_class_overview(self.term_key)
        grade_summary = self.engine.get_class_grade_summary(self.term_key)

        left_items = [
            ("1) Total Number of Students Appeared for Examination", overview.get("total_students", 0)),
            ("2) Total Number of Students Passed in all Subjects (All Clear)", overview.get("passed", 0)),
            ("3) Total Number of Students Failed", overview.get("failed", 0)),
            ("4) Result in % (Without ATKT)", f"{float(overview.get('pass_percent', 0.0)):.2f}%"),
            ("5) Result in % (With ATKT)", f"{float(grade_summary.get('pass_with_atkt', 0.0)):.2f}%"),
        ]

        right_items = [
            ("1) Total Number of Students Appeared for Examination", overview.get("total_students", 0)),
            ("2) Total Number of Students Passed in all Subjects (All Clear)", overview.get("passed", 0)),
            ("3) Total Number of Students Failed", overview.get("failed", 0)),
            ("4) Result in % (Without ATKT)", f"{float(overview.get('pass_percent', 0.0)):.2f}%"),
            ("5) Result in % (With ATKT)", f"{float(grade_summary.get('pass_with_atkt', 0.0)):.2f}%"),
        ]

        start_row = 7
        for idx, (label, value) in enumerate(left_items):
            row = start_row + idx
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(name="Times New Roman", bold=True)
            ws[f"A{row}"].alignment = left
            ws[f"G{row}"] = value
            ws[f"G{row}"].font = Font(name="Times New Roman", bold=True)
            ws[f"G{row}"].alignment = right

        for idx, (label, value) in enumerate(right_items):
            row = start_row + idx
            ws.merge_cells(f"L{row}:R{row}")
            ws[f"L{row}"] = label
            ws[f"L{row}"].font = Font(name="Times New Roman", bold=True)
            ws[f"L{row}"].alignment = left
            ws[f"S{row}"] = value
            ws[f"S{row}"].font = Font(name="Times New Roman", bold=True)
            ws[f"S{row}"].alignment = right

        # Distinction table
        dist_rows = [
            ("DISTINCTION", "> 8"),
            ("FIRST CLASS", "6.5 TO 8"),
            ("SECOND CLASS", "BELOW 6.5"),
        ]
        for i, (label, value) in enumerate(dist_rows):
            row = 6 + i
            ws.merge_cells(f"H{row}:I{row}")
            ws.merge_cells(f"J{row}:K{row}")
            ws[f"H{row}"] = label
            ws[f"J{row}"] = value
            ws[f"H{row}"].font = header_font
            ws[f"J{row}"].font = header_font
            ws[f"H{row}"].alignment = center
            ws[f"J{row}"].alignment = center
            apply_border_range(f"H{row}:K{row}")

        apply_border_range("H6:K8")
        for row in range(6, 9):
            ws.row_dimensions[row].height = 24

        # Summary table
        summary_row = 13
        value_row = summary_row + 1
        summary_layout = [
            ("A", "D", "No of Student Appeared", grade_summary.get("total_students", 0)),
            ("E", "F", "Distinction", grade_summary.get("distinction", 0)),
            ("G", "H", "First class", grade_summary.get("first_class", 0)),
            ("I", "J", "Second class", grade_summary.get("second_class", 0)),
            ("K", "L", "Total Passed", grade_summary.get("passed", 0)),
            ("M", "N", "Passing %", f"{float(grade_summary.get('pass_percent', 0.0)):.2f}%"),
            ("O", "O", "ATKT", grade_summary.get("failed", 0)),
            ("P", "P", "FAILED", grade_summary.get("overall_failed", 0)),
            ("Q", "T", "Passing % With ATKT", f"{float(grade_summary.get('pass_with_atkt', 0.0)):.2f}%"),
            ("U", "U", "BEFORE REMEDIAL", "0"),
        ]

        for start_col, end_col, label, value in summary_layout:
            ws.merge_cells(f"{start_col}{summary_row}:{end_col}{summary_row}")
            cell = ws[f"{start_col}{summary_row}"]
            cell.value = label
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center
            cell.border = border

            ws.merge_cells(f"{start_col}{value_row}:{end_col}{value_row}")
            cell_val = ws[f"{start_col}{value_row}"]
            cell_val.value = value
            cell_val.alignment = center
            cell_val.border = border

            apply_border_range(f"{start_col}{summary_row}:{end_col}{value_row}")

        apply_border_range(f"A{summary_row}:U{summary_row}")
        apply_border_range(f"A{value_row}:U{value_row}")

        # Grade table headers
        grade_header_row = 16
        grade_subheader_row = 17

        ws.merge_cells(f"C{grade_header_row}:L{grade_header_row}")
        ws[f"C{grade_header_row}"] = "GRADE"
        ws[f"C{grade_header_row}"].font = header_font
        ws[f"C{grade_header_row}"].fill = yellow_fill
        ws[f"C{grade_header_row}"].alignment = center
        ws[f"C{grade_header_row}"].border = border

        ws.merge_cells(f"N{grade_header_row}:Q{grade_header_row}")
        ws[f"N{grade_header_row}"] = "BEFORE REMEDIAL."
        ws[f"N{grade_header_row}"].font = header_font
        ws[f"N{grade_header_row}"].fill = yellow_fill
        ws[f"N{grade_header_row}"].alignment = center
        ws[f"N{grade_header_row}"].border = border

        header_pairs = [
            ("A", "SUB"),
            ("B", "NAME OF FACULTY"),
            ("C", "O"),
            ("D", "A++"),
            ("E", "A+"),
            ("F", "A"),
            ("G", "B+"),
            ("H", "B"),
            ("I", "C+"),
            ("J", "C"),
            ("K", "D"),
            ("L", "F/FAIL"),
            ("M", "TOTAL"),
            ("N", "APPEARED"),
            ("O", "PASS"),
            ("P", "FAIL"),
            ("Q", "NOT APPEARED"),
            ("R", "BEFORE REMEDIAL RESULT"),
            ("S", "REM(No of stud passed in remedial exam)"),
            ("T", "AFTER REMEDIAL RESULT"),
            ("U", "SIGN"),
        ]

        for col, label in header_pairs:
            cell = ws[f"{col}{grade_subheader_row}"]
            cell.value = label
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center
            cell.border = border

        for col, label in [("A", "SUB"), ("B", "NAME OF FACULTY")]:
            ws.merge_cells(f"{col}{grade_header_row}:{col}{grade_subheader_row}")
            cell = ws[f"{col}{grade_header_row}"]
            cell.value = label
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center
            cell.border = border

        # Keep S-T-U-V labels on the subheader row (match reference layout).

        # Ensure header rows are fully filled (no white gaps)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=grade_header_row, column=col)
            cell.fill = yellow_fill
            cell.border = border
            if cell.alignment is None:
                cell.alignment = center

        # Column widths
        column_widths = {
            "A": 30,
            "B": 22,
            "C": 6,
            "D": 6,
            "E": 6,
            "F": 6,
            "G": 8,
            "H": 8,
            "I": 8,
            "J": 8,
            "K": 6,
            "L": 7,
            "M": 7,
            "N": 7,
            "O": 9,
            "P": 9,
            "Q": 10,
            "R": 14,
            "S": 14,
            "T": 16,
            "U": 14,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Row heights for cleaner layout
        ws.row_dimensions[6].height = 26
        for r in range(7, 12):
            ws.row_dimensions[r].height = 20
        ws.row_dimensions[7].height = 24
        ws.row_dimensions[8].height = 24
        ws.row_dimensions[summary_row].height = 26
        ws.row_dimensions[value_row].height = 24
        ws.row_dimensions[grade_header_row].height = 26
        ws.row_dimensions[grade_subheader_row].height = 26

        # Grade data rows
        grade_df = self.engine.get_subject_grade_distribution(self.term_key)
        start_data_row = grade_subheader_row + 1
        if not grade_df.empty:
            for idx, row in grade_df.iterrows():
                r = start_data_row + idx
                subject_label = row["subject"]
                ws[f"A{r}"] = subject_label
                ws[f"B{r}"] = row.get("faculty", "")
                ws[f"C{r}"] = int(row.get("O", 0))
                ws[f"D{r}"] = int(row.get("A++", 0))
                ws[f"E{r}"] = int(row.get("A+", 0))
                ws[f"F{r}"] = int(row.get("A", 0))
                ws[f"G{r}"] = int(row.get("B+", 0))
                ws[f"H{r}"] = int(row.get("B", 0))
                ws[f"I{r}"] = int(row.get("C+", 0))
                ws[f"J{r}"] = int(row.get("C", 0))
                ws[f"K{r}"] = int(row.get("D", 0))
                ws[f"L{r}"] = int(row.get("F", 0))
                ws[f"M{r}"] = int(row.get("total", 0))
                ws[f"N{r}"] = int(row.get("appeared", 0))
                ws[f"O{r}"] = int(row.get("passed", 0))
                ws[f"P{r}"] = int(row.get("failed", 0))
                ws[f"Q{r}"] = int(row.get("not_appeared", 0))

                before_remedial = row.get("before_remedial_result")
                ws[f"R{r}"] = "NA" if before_remedial is None else f"{float(before_remedial):.2f}%"
                ws[f"S{r}"] = "Nil"
                ws[f"T{r}"] = "NA"
                ws[f"U{r}"] = ""

                ws.row_dimensions[r].height = 18
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=col)
                    cell.border = border
                    if col in [1, 2]:
                        cell.alignment = left
                    else:
                        cell.alignment = center
        else:
            r = start_data_row
            ws.row_dimensions[r].height = 18
            for col in range(1, max_col + 1):
                cell = ws.cell(row=r, column=col, value="")
                cell.border = border
                cell.alignment = left if col in [1, 2] else center

        # Apply borders to summary and headers
        for r in range(grade_header_row, start_data_row + max(len(grade_df), 1)):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                cell.border = border

    def generate_report(self, output_file):
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        overview = self.engine.get_class_overview(self.term_key)
        toppers_df = self.engine.get_class_toppers(self.top_n, term_key=self.term_key)
        subject_df = self.engine.get_subject_analysis(self.term_key)
        matrix = self.engine.get_student_matrix(self.term_key)
        if subject_df.empty:
            subject_df = pd.DataFrame(columns=["subject", "passed", "failed", "pass_percent"])

        matrix_data = []
        for roll, data in matrix.items():
            matrix_data.append({
                "Roll No": roll,
                "Student Name": data["name"],
                "Passed All Subjects": data["passed_all"],
                "Failed Subjects": ", ".join(data["failed_subjects"])
           })

        matrix_df = pd.DataFrame(matrix_data)

        total_students = int(overview.get("total_students", 0))
        passed = int(overview.get("passed", 0))
        failed = int(overview.get("failed", 0))
        grade_summary = self.engine.get_class_grade_summary(self.term_key)
        overall_failed = int(grade_summary.get("overall_failed", 0))
        report_sgpa_col = self.engine.get_term_sgpa_col(self.term_key) if self.term_key else self.engine.sgpa_col

        summary_subject_rows = []
        for idx, row in subject_df.reset_index(drop=True).iterrows():
            subject_name = str(row.get("subject_name", "")).strip()
            course_code = str(row.get("course_code", "")).strip()
            subject_label = subject_name if not course_code else f"{subject_name} ({course_code})"
            summary_subject_rows.append(
                {
                    "Sr.No": idx + 1,
                    "Subject Teacher": "",
                    "Subject Name": subject_label,
                    "No of Students appeared": int(row.get("appeared", 0)),
                    "No of Students passed": int(row.get("passed", 0)),
                    "No of Students failed": int(row.get("failed", 0)),
                    "Passing Percentage": float(row.get("pass_percent", 0.0)),
                }
            )
        summary_subject_df = pd.DataFrame(summary_subject_rows)

        subject_display = subject_df.rename(
            columns={
                "subject_name": "Subject Name",
                "course_code": "Course Code",
                "subject": "Subject",
                "appeared": "Appeared",
                "passed": "Passed",
                "failed": "Failed",
                "pass_percent": "Pass Percent",
                "topper_name": "Topper Name",
                "topper_prn": "Topper PRN",
                "topper_marks": "Topper Marks",
            }
        )

        toppers_display = toppers_df.rename(
            columns={
                self.engine.name_col: "Student Name",
                self.engine.roll_col: "PRN No",
                report_sgpa_col: "SGPA",
                "grade": "Grade",
            }
        )

        matrix_display = matrix_df.rename(
            columns={
                "Roll No": "PRN No",
                "Student Name": "Student Name",
                "Passed All Subjects": "Passed All Subjects",
                "Failed Subjects": "Failed Subjects",
            }
        )

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            workbook = writer.book
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

            ws = workbook.create_sheet("Summary")
            writer.sheets["Summary"] = ws

            max_col = 7
            end_col = get_column_letter(max_col)

            ws.merge_cells(f"A1:{end_col}1")
            ws["A1"] = self.university
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells(f"A2:{end_col}2")
            ws["A2"] = "School of Engineering & Technology"
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells(f"A3:{end_col}3")
            ws["A3"] = f"Department: {self.department}"
            ws["A3"].alignment = Alignment(horizontal="center")

            ws.merge_cells(f"A4:{end_col}4")
            ws["A4"] = f"A.Y. {self.academic_year}"
            ws["A4"].alignment = Alignment(horizontal="center")

            ws["A6"] = f"Class : {self.class_name or ''}".strip()
            ws["A6"].font = Font(bold=True)
            ws["F6"] = f"Date: {pd.Timestamp.now().strftime('%d/%m/%Y')}"
            ws["F6"].font = Font(bold=True)
            ws["F7"] = f"Semester: {self.semester}"
            ws["F7"].font = Font(bold=True)

            ws["A8"] = f"Total No of Students : {total_students}"
            ws["A9"] = f"Total No of students Appear : {total_students}"
            ws["A10"] = f"Pass : {passed}"
            ws["A11"] = f"ATKT : {failed}"
            ws["A12"] = f"Failed : {overall_failed}"
            ws["A13"] = "Detained : 0"
            ws["A14"] = f"Total Eligible Students for Next year : {passed}"

            for row in range(8, 15):
                ws[f"A{row}"].font = Font(bold=True)

            if not summary_subject_df.empty:
                summary_subject_df.to_excel(
                    writer,
                    sheet_name="Summary",
                    startrow=14,
                    index=False
                )
                ws = writer.sheets["Summary"]
                self.style_table(ws, summary_subject_df, start_row=15)
                self.apply_pass_percent_fill(ws, start_row=16, row_count=len(summary_subject_df), col_idx=7)

            subject_display.to_excel(
                writer,
                sheet_name="Subject Analysis",
                startrow=6,
                index=False
            )

            ws = writer.sheets["Subject Analysis"]
            self.add_header(ws, max_col=len(subject_display.columns))
            self.style_table(ws, subject_display, start_row=7)
            self.add_interpretation_sheet(writer, subject_df)
            self.add_grade_sheet(writer)

            toppers_display.to_excel(
                writer,
                sheet_name="Class Toppers",
                startrow=6,
                index=False
            )

            ws = writer.sheets["Class Toppers"]
            self.add_header(ws, max_col=len(toppers_display.columns))
            self.style_table(ws, toppers_display, start_row=7)

            matrix_display.to_excel(
                writer,
                sheet_name="Student Matrix",
                startrow=6,
                index=False
            )

            ws = writer.sheets["Student Matrix"]
            self.add_header(ws, max_col=len(matrix_display.columns))
            self.style_table(ws, matrix_display, start_row=7)

            # Apply Times New Roman to all sheets
            for sheet in writer.book.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        current = cell.font or Font()
                        cell.font = Font(
                            name="Times New Roman",
                            size=current.size,
                            bold=current.bold,
                            italic=current.italic,
                            underline=current.underline,
                            color=current.color,
                        )
        print(f"\nProfessional Academic Report Generated: {output_file}")
